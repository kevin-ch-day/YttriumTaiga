#!/usr/bin/env python3
"""
export_utils.py
Shared export helpers (CSV/XLSX/JSONL/TXT).
Safe defaults for Red Team ops tooling.
"""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Iterable, Dict, Any, Sequence, List


def write_csv(path: str | Path, headers: Sequence[str], rows: Iterable[Sequence[Any]]) -> None:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    with p.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(list(headers))
        for row in rows:
            writer.writerow(list(row))


def write_jsonl(path: str | Path, records: Iterable[Dict[str, Any]]) -> None:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    with p.open("w", encoding="utf-8") as f:
        for rec in records:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")


def write_json(path: str | Path, payload: Any) -> None:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    with p.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def write_text(path: str | Path, lines: Iterable[str]) -> None:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    with p.open("w", encoding="utf-8") as f:
        for line in lines:
            f.write(f"{line}\n")


def csv_to_xlsx(csv_path: str | Path, xlsx_path: str | Path, *, sheet_name: str | None = None) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    csv_path = Path(csv_path)
    xlsx_path = Path(xlsx_path)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name or csv_path.stem

    with csv_path.open(newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)

    # Header styling
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Autosize columns (capped)
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)


def read_jsonl(path: str | Path) -> List[Dict[str, Any]]:
    p = Path(path)
    records: List[Dict[str, Any]] = []
    with p.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            records.append(json.loads(line))
    return records


def jsonl_to_csv(
    jsonl_path: str | Path,
    csv_path: str | Path,
    *,
    headers: Sequence[str] | None = None,
) -> None:
    records = read_jsonl(jsonl_path)
    if headers is None:
        keys = set()
        for rec in records:
            keys.update(rec.keys())
        headers = sorted(keys)
    rows = [[rec.get(h, "") for h in headers] for rec in records]
    write_csv(csv_path, headers, rows)


def json_to_csv(
    json_path: str | Path,
    csv_path: str | Path,
    *,
    headers: Sequence[str] | None = None,
    records_key: str | None = None,
) -> None:
    p = Path(json_path)
    with p.open("r", encoding="utf-8") as f:
        payload = json.load(f)
    if records_key:
        payload = payload.get(records_key, [])
    if not isinstance(payload, list):
        raise ValueError("JSON payload must be a list of objects")
    records = [r for r in payload if isinstance(r, dict)]
    if headers is None:
        keys = set()
        for rec in records:
            keys.update(rec.keys())
        headers = sorted(keys)
    rows = [[rec.get(h, "") for h in headers] for rec in records]
    write_csv(csv_path, headers, rows)
